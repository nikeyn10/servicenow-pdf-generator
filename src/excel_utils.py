"""
Excel utilities for ServiceNow PDF Generator
Generates Excel summary files for monthly ticket reports
"""

import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_monthly_excel_summary(tickets: List[Dict[str, Any]], month: str, output_dir: str, 
                                shared_attachments: Optional[Dict[str, List[str]]] = None,
                                unique_attachments_count: Optional[int] = None) -> str:
    """
    Create an Excel summary file for monthly tickets
    
    Args:
        tickets: List of ticket dictionaries
        month: Month in YYYY-MM format
        output_dir: Directory to save the Excel file
        shared_attachments: Dictionary of shared attachment filenames and their ticket lists
        unique_attachments_count: Number of unique attachments after deduplication
    
    Returns:
        Path to the created Excel file
    """
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    if wb.active:
        wb.remove(wb.active)
    
    # Create sheets
    summary_sheet = wb.create_sheet("Summary", 0)
    tickets_sheet = wb.create_sheet("Tickets Detail", 1)
    
    if shared_attachments:
        shared_sheet = wb.create_sheet("Shared Attachments", 2)
    
    # Create summary sheet
    _create_summary_sheet(summary_sheet, tickets, month, shared_attachments, unique_attachments_count)
    
    # Create tickets detail sheet
    _create_tickets_sheet(tickets_sheet, tickets)
    
    # Create shared attachments sheet if applicable
    if shared_attachments:
        _create_shared_attachments_sheet(shared_sheet, shared_attachments)
    
    # Save the file
    filename = f"{month}-Resolved-Tickets-Summary.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Ensure output directory exists
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    wb.save(filepath)
    return filepath


def _create_summary_sheet(sheet, tickets: List[Dict[str, Any]], month: str, 
                         shared_attachments: Optional[Dict[str, List[str]]] = None,
                         unique_attachments_count: Optional[int] = None):
    """Create the summary overview sheet"""
    
    # Set sheet title
    sheet.title = "Summary"
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Title
    sheet['A1'] = f"ServiceNow Resolved Tickets Summary - {month}"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet.merge_cells('A1:C1')
    
    # Generation info
    sheet['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sheet['A3'].font = Font(italic=True)
    
    # Summary statistics
    row = 5
    
    # Basic stats
    stats = [
        ("Total Resolved Tickets", len(tickets)),
        ("Total Attachments", sum(len(ticket.get('attachments', [])) for ticket in tickets)),
    ]
    
    if unique_attachments_count is not None:
        stats.append(("Unique Attachments (After Deduplication)", unique_attachments_count))
        stats.append(("Duplicate Attachments Removed", sum(len(ticket.get('attachments', [])) for ticket in tickets) - unique_attachments_count))
    
    if shared_attachments:
        stats.append(("Shared Attachment Files", len(shared_attachments)))
    
    # Add tickets with attachments count
    tickets_with_attachments = len([t for t in tickets if t.get('attachments')])
    stats.append(("Tickets with Attachments", tickets_with_attachments))
    stats.append(("Tickets without Attachments", len(tickets) - tickets_with_attachments))
    
    # Write stats
    for stat_name, stat_value in stats:
        sheet[f'A{row}'] = stat_name
        sheet[f'B{row}'] = stat_value
        sheet[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Status breakdown if available
    row += 2
    sheet[f'A{row}'] = "Ticket Status Breakdown"
    sheet[f'A{row}'].font = Font(bold=True, size=12)
    sheet[f'A{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    row += 1
    
    # Count by status
    status_counts = {}
    for ticket in tickets:
        status = ticket.get('status', 'Unknown')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    for status, count in status_counts.items():
        sheet[f'A{row}'] = status
        sheet[f'B{row}'] = count
        row += 1
    
    # Date range analysis
    row += 2
    sheet[f'A{row}'] = "Date Analysis"
    sheet[f'A{row}'].font = Font(bold=True, size=12)
    sheet[f'A{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    row += 1
    
    # Extract dates and find range
    dates = []
    for ticket in tickets:
        date_str = ticket.get('date_opened')
        if date_str:
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                dates.append(date_obj)
            except:
                pass
    
    if dates:
        earliest = min(dates)
        latest = max(dates)
        sheet[f'A{row}'] = "Earliest Ticket Date"
        sheet[f'B{row}'] = earliest.strftime('%Y-%m-%d')
        row += 1
        sheet[f'A{row}'] = "Latest Ticket Date"
        sheet[f'B{row}'] = latest.strftime('%Y-%m-%d')
        row += 1
    
    # Auto-fit columns
    for col in ['A', 'B', 'C']:
        sheet.column_dimensions[col].auto_size = True


def _create_tickets_sheet(sheet, tickets: List[Dict[str, Any]]):
    """Create the detailed tickets listing sheet"""
    
    sheet.title = "Tickets Detail"
    
    # Headers
    headers = [
        "Ticket #", "Item ID", "Date Opened", "Status", "Summary", 
        "Attachments Count", "Attachment Files", "Has Shared Attachments"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write ticket data
    for row, ticket in enumerate(tickets, 2):
        sheet.cell(row=row, column=1, value=ticket.get('name', ''))
        sheet.cell(row=row, column=2, value=ticket.get('id', ''))
        sheet.cell(row=row, column=3, value=ticket.get('date_opened', ''))
        sheet.cell(row=row, column=4, value=ticket.get('status', ''))
        
        # Summary (truncated for Excel)
        summary = ticket.get('summary', '')
        if len(summary) > 100:
            summary = summary[:97] + "..."
        sheet.cell(row=row, column=5, value=summary)
        
        # Attachments info
        attachments = ticket.get('attachments', [])
        sheet.cell(row=row, column=6, value=len(attachments))
        
        # Attachment files (first few filenames)
        if attachments:
            file_list = ', '.join(attachments[:3])  # Show first 3 files
            if len(attachments) > 3:
                file_list += f" ... (+{len(attachments) - 3} more)"
            sheet.cell(row=row, column=7, value=file_list)
        else:
            sheet.cell(row=row, column=7, value="No attachments")
        
        # Has shared attachments indicator
        # This would need to be calculated based on shared_attachments data
        sheet.cell(row=row, column=8, value="TBD")  # Placeholder
    
    # Create table
    if len(tickets) > 0:
        table_range = f"A1:{get_column_letter(len(headers))}{len(tickets) + 1}"
        table = Table(displayName="TicketsTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        sheet.add_table(table)
    
    # Auto-fit columns
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        sheet.column_dimensions[column_letter].width = adjusted_width


def _create_shared_attachments_sheet(sheet, shared_attachments: Dict[str, List[str]]):
    """Create the shared attachments analysis sheet"""
    
    sheet.title = "Shared Attachments"
    
    # Headers
    headers = ["Attachment Filename", "Usage Count", "Tickets Using This File"]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write shared attachment data
    row = 2
    for filename, ticket_list in shared_attachments.items():
        sheet.cell(row=row, column=1, value=filename)
        sheet.cell(row=row, column=2, value=len(ticket_list))
        sheet.cell(row=row, column=3, value=', '.join(ticket_list))
        row += 1
    
    # Create table if we have data
    if shared_attachments:
        table_range = f"A1:C{row - 1}"
        table = Table(displayName="SharedAttachmentsTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium12", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        sheet.add_table(table)
    
    # Auto-fit columns
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)  # Cap at 80 characters for ticket lists
        sheet.column_dimensions[column_letter].width = adjusted_width


def create_attachment_analysis_excel(tickets: List[Dict[str, Any]], month: str, output_dir: str) -> str:
    """
    Create a detailed Excel analysis of attachments
    
    Args:
        tickets: List of ticket dictionaries
        month: Month in YYYY-MM format
        output_dir: Directory to save the Excel file
    
    Returns:
        Path to the created Excel file
    """
    
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    
    # Create analysis sheet
    analysis_sheet = wb.create_sheet("Attachment Analysis", 0)
    
    # Analyze all attachments
    attachment_usage = {}
    ticket_attachments = {}
    
    for ticket in tickets:
        ticket_id = ticket.get('name', ticket.get('id', ''))
        attachments = ticket.get('attachments', [])
        ticket_attachments[ticket_id] = attachments
        
        for attachment in attachments:
            if attachment not in attachment_usage:
                attachment_usage[attachment] = []
            attachment_usage[attachment].append(ticket_id)
    
    # Create analysis
    _create_attachment_analysis_sheet(analysis_sheet, attachment_usage, month)
    
    # Save the file
    filename = f"{month}-Attachment-Analysis.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    return filepath


def _create_attachment_analysis_sheet(sheet, attachment_usage: Dict[str, List[str]], month: str):
    """Create attachment analysis sheet"""
    
    sheet.title = "Attachment Analysis"
    
    # Title
    sheet['A1'] = f"Attachment Analysis - {month}"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet.merge_cells('A1:D1')
    
    # Headers
    headers = ["Attachment Filename", "Usage Count", "Status", "Tickets"]
    row = 3
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Sort by usage count (most shared first)
    sorted_attachments = sorted(attachment_usage.items(), key=lambda x: len(x[1]), reverse=True)
    
    row = 4
    for filename, tickets in sorted_attachments:
        usage_count = len(tickets)
        status = "Shared" if usage_count > 1 else "Unique"
        
        sheet.cell(row=row, column=1, value=filename)
        sheet.cell(row=row, column=2, value=usage_count)
        sheet.cell(row=row, column=3, value=status)
        sheet.cell(row=row, column=4, value=', '.join(tickets))
        
        # Color code shared attachments
        if usage_count > 1:
            for col in range(1, 5):
                sheet.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFEEEE", end_color="FFEEEE", fill_type="solid"
                )
        
        row += 1
    
    # Auto-fit columns
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        sheet.column_dimensions[column_letter].width = adjusted_width
